import fpdf
import openpyxl
import PyPDF2
from PyPDF2 import PdfFileWriter, PdfFileReader
import os
import tkinter as tk
from tkinter import filedialog
import xlwings as xw
import xlsxwriter
from itertools import groupby

def load_excel():
    try:
        # getting path to CSV file
        global csvpath
        csvpath = filedialog.askopenfile(filetypes = [("CSV files", "*.csv")])
        ex_info.set(csvpath.name)
        label_message.set("")
    except AttributeError:
        ex_info.set("")
        label_message.set("File not loaded")

def load_pdf():
    try:
        # getting path to PDF file
        global pdfpath
        pdfpath = filedialog.askopenfile(filetypes = [("PDF files", "*.pdf")])
        pdf_info.set(pdfpath.name)
        label_message.set("")
    except AttributeError:
        label_message.set("File not loaded")
    except NameError:
        label_message.set("File not loaded")

def run_macro():
    try:
        csvpath.name
        # creating temp.xlsm file, adding macros from vbaProject.bin
        workbook = xlsxwriter.Workbook("temp.xlsm")
        workbook.add_vba_project("vbaProject.bin")
        workbook.close()
        # running macro depending on user's choice
        if check_var.get() == 0:
            macro1()
        else:
            macro2()
    except AttributeError:
        label_message.set("Load CSV file.")

def macro1():
    # copying data from CSV file to yPDF.xlsm
    projectNum = project_entry.get()
    wb = xw.Book("temp.xlsm")
    app = xw.apps.active
    # running macro, saving and closing excel window
    m1 = wb.macro("Module1.macro1")
    m1(projectNum, csvpath.name)
    wb.save()
    app.quit()

def macro2():
    # copying data from CSV file to yPDF.xlsm
    projectNum = project_entry.get()
    wb = xw.Book("temp.xlsm")
    app = xw.apps.active
    # running macro
    m2 = wb.macro("Module2.macro2")
    m2(projectNum, csvpath.name)
    wb.save()
    app.quit()
    os.system("start EXCEL.EXE temp.xlsm")

def ranges(i):
    # grouping function
    for a, b in groupby(enumerate(i), lambda pair: pair[1] - pair[0]):
        b = list(b)
        yield b[0][1], b[-1][1]

def write_data_to_pdf():
    label_message.set("")
    try:
        # loading temp.xlsm of temp_b.xlsm file
        if check_var.get() == 0:
            wb = openpyxl.load_workbook("temp.xlsm")
        else:
            wb = openpyxl.load_workbook("temp_b.xlsm")
        sheet = wb[wb.sheetnames[0]]
        num_g = sheet.max_row

        # creating dictionary with data from excel, deleting useless text
        pairs = {}
        
        for i in range(1,num_g+1):
            if sheet[f"A{i}"].value not in pairs:
                pairs[sheet[f"A{i}"].value] = (sheet[f"B{i}"].value)
            elif isinstance(pairs[sheet[f"A{i}"].value], list):
                pairs[sheet[f"A{i}"].value].append(sheet[f"B{i}"].value)
                # deleting duplicates in each list
                pairs[sheet[f"A{i}"].value] = list(dict.fromkeys(pairs[sheet[f"A{i}"].value]))
                # sorting each list
                pairs[sheet[f"A{i}"].value].sort()
            else:
                pairs[sheet[f"A{i}"].value] = [pairs[sheet[f"A{i}"].value], (sheet[f"B{i}"].value)]

        # creating list of page numbers (keys) to determine max page number
        keys = []
        for i in pairs.keys():
            keys.append(int(i))
        max_page = max(keys)

        # creating dictionary with page sizes
        existing_pdf = PdfFileReader(open(pdfpath.name, "rb"))
        count = existing_pdf.numPages
        sizes = {}
        for i in range(count):
            x = existing_pdf.getPage(i)
            sizes[i+1] = [int(x["/MediaBox"][2]), int(x["/MediaBox"][3])]

        # setting global properties of temporary PDF file and font
        temp_pdf = fpdf.FPDF(orientation = "P", unit = "pt")
        temp_pdf.set_font("Courier")
        temp_pdf.set_margins(0, 0, 0)


        doc_num = doc_entry.get()

        # iterating through dictionary and adding blank or filled pages to temporary PDF file
        for i in range(0,max_page):
            page_w = sizes[i+1][0]
            page_h = sizes[i+1][1]
            cell_w = 0.18
            space_left = 0.08
            if pairs.get(i+1) == None:
                temp_pdf.add_page(format = (page_w, page_h))
            elif type(pairs[i+1]) is str:
                temp_pdf.set_font("", "B", size=int(page_h/33.1))
                temp_pdf.set_text_color(255, 0, 0)
                temp_pdf.set_draw_color(255, 0, 0)
                temp_pdf.set_fill_color(255, 255, 255)
                temp_pdf.set_line_width(3)
                temp_pdf.add_page(format = (page_w, page_h))
                temp_pdf.cell(int((1-space_left-cell_w)*page_w), int((page_h)/36), ln=0)
                temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str(pairs[i+1]), ln=0, border=1, fill=1, align="C")
                if doc_num == "":
                    pass
                else:
                    temp_pdf.cell(int(0.2*space_left*page_w), int((page_h)/36), ln=0)
                    temp_pdf.set_font("", "B", size=int(page_h/50))
                    temp_pdf.set_text_color(26, 26, 255)
                    temp_pdf.set_draw_color(26, 26, 255)
                    temp_pdf.set_line_width(2)
                    temp_pdf.cell(int(0.8*space_left*page_w), int((page_h)/45), txt=f"DOK-{doc_num}", ln=0, border=1, fill=1, align="C")
            else:
                temp_pdf.set_font("", "B", size=int(page_h/33.1))
                temp_pdf.set_text_color(255, 0, 0)
                temp_pdf.set_draw_color(255, 0, 0)
                temp_pdf.set_fill_color(255, 255, 255)
                temp_pdf.set_line_width(3)
                temp_pdf.add_page(format = (page_w, page_h))
                if len(pairs[i+1]) == 1:
                    temp_pdf.cell(int((1-space_left-(cell_w))*page_w), int((page_h)/36), ln=0)
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[0]), ln=0, border=1, fill=1, align="C")
                elif len(pairs[i+1]) == 2:
                    temp_pdf.cell(int((1-space_left-(2*cell_w))*page_w), int((page_h)/36), ln=0)
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[0]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[1]), ln=0, border=1, fill=1, align="C")
                elif len(pairs[i+1]) == 3:
                    temp_pdf.cell(int((1-space_left-(3*cell_w))*(page_w)), int((page_h)/36), ln=0)
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[0]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[1]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[2]), ln=0, border=1, fill=1, align="C")
                elif len(pairs[i+1]) == 4:
                    temp_pdf.cell(int((1-space_left-(4*cell_w))*(page_w)), int((page_h)/36), ln=0)
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[0]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[1]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[2]), ln=0, border=1, fill=1, align="C")
                    temp_pdf.cell(int(cell_w*page_w), int((page_h)/36), txt=str((pairs[i+1])[3]), ln=0, border=1, fill=1, align="C")
                if doc_num == "":
                    pass
                else:
                    temp_pdf.cell(int(0.2*space_left*page_w), int((page_h)/36), ln=0)
                    temp_pdf.set_font("", "B", size=int(page_h/50))
                    temp_pdf.set_text_color(26, 26, 255)
                    temp_pdf.set_draw_color(26, 26, 255)
                    temp_pdf.set_line_width(2)
                    temp_pdf.cell(int(0.8*space_left*page_w), int((page_h)/45), txt=f"DOK-{doc_num}", ln=0, border=1, fill=1, align="C")

        # saving temporary PDF file
        temp_pdf.output("temp.pdf")

        # opening existing file, counting number of pages
        existing_pdf = PdfFileReader(open(pdfpath.name, "rb"))
        with open("temp.pdf", "rb") as f:
            temp = PdfFileReader(f)
            count = existing_pdf.numPages
            output = PdfFileWriter()

            # iterating through pages, merging pages from two PDF files
            if count == max_page:
                for i in range(0,count):
                    page = existing_pdf.getPage(i)
                    page.mergePage(temp.getPage(i))
                    output.addPage(page)
            else:
                for i in range(0,max_page):
                    page = existing_pdf.getPage(i)
                    page.mergePage(temp.getPage(i))
                    output.addPage(page)
                for i in range(max_page, count):
                    page = existing_pdf.getPage(i)
                    output.addPage(page)

            # saving output file
            name = os.path.splitext(os.path.basename(pdfpath.name))[0]
            with open(f"{name}-desc.pdf", "wb") as outputStream:
                output.write(outputStream)

        # deleting temporary files
        os.remove("temp.pdf")
        os.remove("temp.xlsm")
        if check_var.get() == 1:
            os.remove("temp_b.xlsm")
        label_message.set(f"File {name}-desc.pdf was created successfully!")

        # creating plot.txt file
        grouped_keys = list(ranges(keys))
        plot_list = []
        for i in range(len(grouped_keys)):
            if grouped_keys[i][0] == grouped_keys[i][1]:
                plot_list.append(f"{grouped_keys[i][0]}")
            else:
                plot_list.append(f"{grouped_keys[i][0]}-{grouped_keys[i][1]}")
        plot_text = (",".join(map(str,plot_list)))

        with open("plot.txt", "w") as plot:
            plot.write(plot_text)
    
    # handling errors
    except ValueError:
        os.remove("temp.xlsm")
        label_message.set("Insert correct project number.")
    except NameError:
        label_message.set("At least one of the files is not loaded.")
    except openpyxl.utils.exceptions.InvalidFileException:
        label_message.set("Load CSV file.")
    except xlsxwriter.exceptions.FileCreateError:
        label_message.set("Load CSV file.")
    except PyPDF2.utils.PdfReadError:
        temp_pdf.close()
        os.remove("temp.pdf")
        label_message.set("Load PDF file.")
    except PermissionError:
        temp_pdf.close()
        os.remove("temp.pdf")
        label_message.set(f"File {name}-desc.pdf is open. Cannot overwrite.")
    except:
        label_message.set("Something went wrong, excel file may contain wrong page numbers. Check the 'excel correction' checkbox.")

def do_everything():
    run_macro()
    write_data_to_pdf()

root = tk.Tk()
root.geometry("730x170")
root.title("yPDF")
root.resizable(False, False)

label_message = tk.StringVar()
ex_info = tk.StringVar()
pdf_info = tk.StringVar()
check_var = tk.IntVar()

FONT = ("Arial", 10)

frame1 = tk.Frame(root, width=500, height=28)
frame1.pack()
frame1.pack_propagate(0)


label_num = tk.Label(frame1, text="Project number:", font=FONT, width=12)
label_num.grid(row=0, column=0)

project_entry = tk.Entry(frame1, font=FONT, width=12)
project_entry.grid(row=0, column=1, padx=(0,35))

label_doc = tk.Label(frame1, text="Doc number:", font=FONT, width=10)
label_doc.grid(row=0, column=2)

doc_entry = tk.Entry(frame1, font=FONT, width=5)
doc_entry.grid(row=0, column=3, padx=(0,45))

ch_box = tk.Checkbutton(frame1, text="change something in excel", variable=check_var)
ch_box.grid(row=0, column=4, padx=(0,30))

frame2 = tk.Frame(root)
frame2.pack(pady=5)

ex_button = tk.Button(frame2, text="Load CSV file", font=FONT, command=load_excel, width=20)
ex_button.grid(row=1, column=0, padx=5)

label_ex = tk.Label(frame2, textvariable=ex_info, font=FONT, borderwidth=2, relief="groove", width=62, height=1)
label_ex.grid(row=1, column=1, padx=5)

pdf_button = tk.Button(frame2, text="Load PDF file", font=FONT, command=load_pdf, width=20)
pdf_button.grid(row=2, column=0, padx=5)

label_pdf = tk.Label(frame2, textvariable=pdf_info, font=FONT, borderwidth=2, relief="groove", width=62, height=1)
label_pdf.grid(row=2, column=1, padx=5)

frame3 = tk.Frame(root)
frame3.pack()

macro_button = tk.Button(frame3, text="Run macro", font=FONT, command=run_macro, width=20)
macro_button.grid(row=0, column=0)

data_button = tk.Button(frame3, text="Write data to PDF", font=FONT, command=write_data_to_pdf, width=20)
data_button.grid(row=0, column=1)

ev_button = tk.Button(frame3, text="Do everything", font=FONT, command=do_everything, width=20)
ev_button.grid(row=0, column=2)

frame4 = tk.Frame(root)
frame4.pack(pady=(5,0))

label_info = tk.Label(frame4, textvariable=label_message, font=("Arial", 10, "bold"), width=80, height=2, wraplength=500)
label_info.grid(row=0, column=0)

root.mainloop()